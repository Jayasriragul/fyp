import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


HEADER_FONT = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='6C3CE1', end_color='6C3CE1', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center')
CELL_FONT = Font(name='Calibri', size=11)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def _style_header(ws, headers, row=1):
    """Apply header styling to worksheet."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _auto_width(ws):
    """Auto-adjust column widths."""
    for column_cells in ws.columns:
        length = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 4, 50)


def export_users_excel(users):
    """Export users list to Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Users'

    headers = ['ID', 'Name', 'Email', 'Phone', 'Role', 'Blacklisted', 'Created At']
    _style_header(ws, headers)

    for row_idx, user in enumerate(users, 2):
        ws.cell(row=row_idx, column=1, value=user.id).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=user.name).border = THIN_BORDER
        ws.cell(row=row_idx, column=3, value=user.email).border = THIN_BORDER
        ws.cell(row=row_idx, column=4, value=user.phone or '').border = THIN_BORDER
        ws.cell(row=row_idx, column=5, value=user.role).border = THIN_BORDER
        ws.cell(row=row_idx, column=6, value='Yes' if user.is_blacklisted else 'No').border = THIN_BORDER
        ws.cell(row=row_idx, column=7, value=user.created_at.strftime('%Y-%m-%d %H:%M') if user.created_at else '').border = THIN_BORDER

    _auto_width(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_attendance_excel(attendance_records):
    """Export attendance records to Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance'

    headers = ['ID', 'User Name', 'Email', 'Event', 'Check In', 'Check Out', 'Type', 'Day']
    _style_header(ws, headers)

    for row_idx, record in enumerate(attendance_records, 2):
        user = record.registration.user if record.registration else None
        event = record.event
        ws.cell(row=row_idx, column=1, value=record.id).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=user.name if user else '').border = THIN_BORDER
        ws.cell(row=row_idx, column=3, value=user.email if user else '').border = THIN_BORDER
        ws.cell(row=row_idx, column=4, value=event.title if event else '').border = THIN_BORDER
        ws.cell(row=row_idx, column=5, value=record.check_in_time.strftime('%Y-%m-%d %H:%M') if record.check_in_time else '').border = THIN_BORDER
        ws.cell(row=row_idx, column=6, value=record.check_out_time.strftime('%Y-%m-%d %H:%M') if record.check_out_time else '').border = THIN_BORDER
        ws.cell(row=row_idx, column=7, value=record.scan_type).border = THIN_BORDER
        ws.cell(row=row_idx, column=8, value=record.day_number).border = THIN_BORDER

    _auto_width(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_meals_excel(meal_records):
    """Export meal records to Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Meals'

    headers = ['ID', 'User Name', 'Email', 'Event', 'Meal Type', 'Scanned At', 'Day']
    _style_header(ws, headers)

    for row_idx, record in enumerate(meal_records, 2):
        user = record.registration.user if record.registration else None
        event = record.event
        ws.cell(row=row_idx, column=1, value=record.id).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=user.name if user else '').border = THIN_BORDER
        ws.cell(row=row_idx, column=3, value=user.email if user else '').border = THIN_BORDER
        ws.cell(row=row_idx, column=4, value=event.title if event else '').border = THIN_BORDER
        ws.cell(row=row_idx, column=5, value=record.meal_type).border = THIN_BORDER
        ws.cell(row=row_idx, column=6, value=record.scanned_at.strftime('%Y-%m-%d %H:%M') if record.scanned_at else '').border = THIN_BORDER
        ws.cell(row=row_idx, column=7, value=record.day_number).border = THIN_BORDER

    _auto_width(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_feedback_excel(feedback_records):
    """Export feedback records to Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Feedback'

    headers = ['ID', 'User Name', 'Email', 'Event', 'Rating', 'Comment', 'Sentiment', 'Date']
    _style_header(ws, headers)

    for row_idx, record in enumerate(feedback_records, 2):
        user = record.user
        event = record.event
        ws.cell(row=row_idx, column=1, value=record.id).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=user.name if user else '').border = THIN_BORDER
        ws.cell(row=row_idx, column=3, value=user.email if user else '').border = THIN_BORDER
        ws.cell(row=row_idx, column=4, value=event.title if event else '').border = THIN_BORDER
        ws.cell(row=row_idx, column=5, value=record.rating).border = THIN_BORDER
        ws.cell(row=row_idx, column=6, value=record.comment or '').border = THIN_BORDER
        ws.cell(row=row_idx, column=7, value=record.sentiment).border = THIN_BORDER
        ws.cell(row=row_idx, column=8, value=record.created_at.strftime('%Y-%m-%d %H:%M') if record.created_at else '').border = THIN_BORDER

    _auto_width(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
